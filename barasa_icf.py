import os
import tempfile
from barcode import Code128
from barcode.writer import ImageWriter
from openpyxl import Workbook
from openpyxl.drawing.image import Image

# Define the range for the barcodes
start_code = 1
end_code = 350
copies_per_barcode = 1  # Number of copies per barcode

# Create a temporary directory to store the barcode images
temp_dir = tempfile.mkdtemp()

# Initialize the Excel workbook and worksheet
wb = Workbook()
ws = wb.active
ws.title = "Barcodes"

# Add headers
ws.append(["Barcode Data", "Copy Number", "Barcode Image"])

# Generate the barcodes and add them to the Excel sheet
row_index = 2  # Start from row 2 (row 1 is for headers)

for code in range(start_code, end_code + 1):
    base_data = f"{code:03d}-CRF"  # → 001-ICF format

    for copy in range(1, copies_per_barcode + 1):
        # Generate the barcode image
        barcode = Code128(base_data, writer=ImageWriter())
        temp_image_path = os.path.join(temp_dir, f"barcode_{code}_{copy}.png")
        barcode.save(os.path.splitext(temp_image_path)[0])

        # Write text data
        ws.cell(row=row_index, column=1, value=base_data)
        ws.cell(row=row_index, column=2, value=copy)

        # Insert the barcode image
        img = Image(temp_image_path)
        img.width, img.height = 200, 100
        ws.add_image(img, f"C{row_index}")

        row_index += 1

# Save file
excel_file = "CRF_Barcodes.xlsx"
wb.save(excel_file)

# Cleanup
for file_name in os.listdir(temp_dir):
    file_path = os.path.join(temp_dir, file_name)
    if os.path.isfile(file_path):
        os.remove(file_path)
os.rmdir(temp_dir)

print(f"All barcodes generated and saved in '{excel_file}'.")
