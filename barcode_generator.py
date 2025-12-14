from openpyxl import Workbook


def generate_barcodes(
    start_code=1,
    end_code=350,
    copies_per_barcode=1,
    barcode_format="{code:03d}-ICF",
    output_file="barcodes.xlsx"
):
    """
    Generate barcodes with customizable attributes and save to Excel.

    Parameters:
    -----------
    start_code : int
        Starting number for barcode sequence (default: 1)
    end_code : int
        Ending number for barcode sequence (default: 350)
    copies_per_barcode : int
        Number of copies for each barcode (default: 1)
    barcode_format : str
        Format string for barcode. Use {code} as placeholder (default: "{code:03d}-ICF")
        Examples:
            - "{code:03d}-ICF" → 001-ICF, 002-ICF, ...
            - "{code:03d}-CRF" → 001-CRF, 002-CRF, ...
            - "BAR-{code:04d}" → BAR-0001, BAR-0002, ...
    output_file : str
        Name of output Excel file (default: "barcodes.xlsx")

    Returns:
    --------
    str : Path to the generated Excel file
    """
    # Initialize the Excel workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Barcodes"

    # Add headers
    ws.append(["Barcode Data", "Copy Number"])

    # Generate the barcode data and add to Excel sheet
    row_index = 2  # Start from row 2 (row 1 is for headers)

    for code in range(start_code, end_code + 1):
        barcode_data = barcode_format.format(code=code)

        for copy in range(1, copies_per_barcode + 1):
            ws.cell(row=row_index, column=1, value=barcode_data)
            ws.cell(row=row_index, column=2, value=copy)
            row_index += 1

    # Save the Excel file
    wb.save(output_file)

    print(f"Barcodes generated successfully!")
    print(f"Range: {start_code} to {end_code}")
    print(f"Copies per barcode: {copies_per_barcode}")
    print(f"Total rows: {row_index - 2}")
    print(f"Saved to: '{output_file}'")

    return output_file


if __name__ == "__main__":
    # Example usage - customize these parameters as needed
    generate_barcodes(
        start_code=1,
        end_code=350,
        copies_per_barcode=1,
        barcode_format="{code:03d}-ICF",
        output_file="ICF_Barcodes.xlsx"
    )
